"""
Tests for the excel_config module.

Tests cover:
- Variable role heuristic classification
- Display name generation
- MFF data discovery
- Template generation (MFF → Excel)
- Template parsing (Excel → configs)
- Round-trip: generate → parse → verify
- Validation error cases
"""

from pathlib import Path
import tempfile

import numpy as np
import pandas as pd
import pytest

from mmm_framework.excel_config import (
    DiscoveredVariable,
    TemplateGenerator,
    TemplateParser,
    TemplateParseError,
    TemplateValidationError,
    VariableStats,
    classify_variable,
    discover_mff,
    generate_display_name,
)
from mmm_framework.config import (
    AdstockType,
    DimensionType,
    InferenceMethod,
    ModelSpecification,
    SaturationType,
    VariableRole,
)


# =============================================================================
# Fixtures
# =============================================================================


def _create_sample_mff(
    n_periods: int = 52,
    geographies: list[str] | None = None,
    products: list[str] | None = None,
) -> pd.DataFrame:
    """Create a sample MFF DataFrame for testing."""
    if geographies is None:
        geographies = ["North", "South", "East", "West"]
    if products is None:
        products = ["Product_A"]

    rows = []
    dates = pd.date_range("2024-01-01", periods=n_periods, freq="W")

    variables = {
        "Sales": {"role": "kpi", "base": 10000, "noise": 2000},
        "TV_Spend": {"role": "media", "base": 5000, "noise": 3000},
        "Digital_Spend": {"role": "media", "base": 3000, "noise": 2000},
        "Radio_Spend": {"role": "media", "base": 1000, "noise": 500},
        "Price": {"role": "control", "base": 9.99, "noise": 1.0},
        "Holiday": {"role": "control", "base": 0, "noise": 0.3},
        "Distribution": {"role": "control", "base": 80, "noise": 5},
    }

    rng = np.random.default_rng(42)

    for date in dates:
        for geo in geographies:
            for product in products:
                for var_name, var_info in variables.items():
                    value = max(0, var_info["base"] + rng.normal(0, var_info["noise"]))
                    if var_name == "Holiday":
                        value = 1.0 if rng.random() < 0.1 else 0.0

                    rows.append({
                        "Period": date.strftime("%Y-%m-%d"),
                        "Geography": geo,
                        "Product": product,
                        "Campaign": "All",
                        "Outlet": "All",
                        "Creative": "All",
                        "VariableName": var_name,
                        "VariableValue": round(value, 2),
                    })

    return pd.DataFrame(rows)


@pytest.fixture
def sample_mff_df() -> pd.DataFrame:
    """Sample MFF DataFrame with geo-level data."""
    return _create_sample_mff()


@pytest.fixture
def national_mff_df() -> pd.DataFrame:
    """Sample MFF DataFrame with national-only data."""
    return _create_sample_mff(geographies=["National"], products=["All"])


@pytest.fixture
def geo_product_mff_df() -> pd.DataFrame:
    """Sample MFF with both geo and product dimensions."""
    return _create_sample_mff(
        geographies=["North", "South"],
        products=["Product_A", "Product_B"],
    )


@pytest.fixture
def tmp_dir():
    """Temporary directory for test files."""
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


# =============================================================================
# Heuristic Classification Tests
# =============================================================================


class TestClassifyVariable:
    """Tests for variable role classification heuristics."""

    def test_kpi_patterns(self):
        """Common KPI variable names should be classified as KPI."""
        assert classify_variable("Sales") == VariableRole.KPI
        assert classify_variable("Revenue") == VariableRole.KPI
        assert classify_variable("Conversions") == VariableRole.KPI
        assert classify_variable("Total_Orders") == VariableRole.KPI
        assert classify_variable("Leads") == VariableRole.KPI
        assert classify_variable("Downloads") == VariableRole.KPI
        assert classify_variable("Signups") == VariableRole.KPI

    def test_media_patterns(self):
        """Common media variable names should be classified as Media."""
        assert classify_variable("TV_Spend") == VariableRole.MEDIA
        assert classify_variable("Digital_Impressions") == VariableRole.MEDIA
        assert classify_variable("Radio_GRPs") == VariableRole.MEDIA
        assert classify_variable("Facebook_Spend") == VariableRole.MEDIA
        assert classify_variable("Google_Clicks") == VariableRole.MEDIA
        assert classify_variable("YouTube_Views") == VariableRole.MEDIA
        assert classify_variable("TikTok_Impressions") == VariableRole.MEDIA
        assert classify_variable("OOH_Spend") == VariableRole.MEDIA
        assert classify_variable("Programmatic_Display") == VariableRole.MEDIA

    def test_control_patterns(self):
        """Common control variable names should be classified as Control."""
        assert classify_variable("Price") == VariableRole.CONTROL
        assert classify_variable("Promotion_Flag") == VariableRole.CONTROL
        assert classify_variable("Distribution") == VariableRole.CONTROL
        assert classify_variable("Holiday") == VariableRole.CONTROL
        assert classify_variable("Temperature") == VariableRole.CONTROL
        assert classify_variable("Competitor_Price") == VariableRole.CONTROL
        assert classify_variable("GDP_Index") == VariableRole.CONTROL

    def test_unknown_defaults_to_auxiliary(self):
        """Unknown variable names should default to AUXILIARY (Exclude)."""
        assert classify_variable("XYZ123") == VariableRole.AUXILIARY
        assert classify_variable("FooBar") == VariableRole.AUXILIARY

    def test_case_insensitive(self):
        """Classification should be case-insensitive."""
        assert classify_variable("sales") == VariableRole.KPI
        assert classify_variable("SALES") == VariableRole.KPI
        assert classify_variable("tv_spend") == VariableRole.MEDIA
        assert classify_variable("TV_SPEND") == VariableRole.MEDIA

    def test_statistical_fallback(self):
        """Statistical properties should help classify when keywords fail."""
        # High zero percentage + high variance → media-like
        media_stats = VariableStats(
            mean=1000, std=2000, min_val=0, max_val=10000,
            zero_pct=0.5, n_obs=100, coverage_pct=95.0,
        )
        result = classify_variable("Channel_X", stats=media_stats)
        assert result == VariableRole.MEDIA

        # Low variance → control-like
        control_stats = VariableStats(
            mean=100, std=2, min_val=95, max_val=105,
            zero_pct=0.0, n_obs=100, coverage_pct=100.0,
        )
        result = classify_variable("Factor_Y", stats=control_stats)
        assert result == VariableRole.CONTROL


class TestGenerateDisplayName:
    """Tests for display name generation."""

    def test_snake_case(self):
        assert generate_display_name("tv_spend") == "TV Spend"
        assert generate_display_name("digital_impressions") == "Digital Impressions"

    def test_camel_case(self):
        assert generate_display_name("PaidSearch") == "Paid Search"
        assert generate_display_name("TotalSales") == "Total Sales"

    def test_acronyms_preserved(self):
        assert generate_display_name("gdp") == "GDP"
        assert generate_display_name("cpi_index") == "CPI Index"
        assert generate_display_name("tv") == "TV"

    def test_already_clean(self):
        assert generate_display_name("Price") == "Price"


# =============================================================================
# Discovery Tests
# =============================================================================


class TestDiscoverMFF:
    """Tests for MFF data discovery."""

    def test_discovers_all_variables(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        names = {v.name for v in discovery.variables}
        assert "Sales" in names
        assert "TV_Spend" in names
        assert "Price" in names
        assert len(discovery.variables) == 7

    def test_discovers_geographies(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        assert set(discovery.geographies) == {"East", "North", "South", "West"}

    def test_discovers_national(self, national_mff_df):
        discovery = discover_mff(national_mff_df)
        assert discovery.geographies == []  # "National" is filtered out

    def test_discovers_products(self, geo_product_mff_df):
        discovery = discover_mff(geo_product_mff_df)
        assert set(discovery.products) == {"Product_A", "Product_B"}

    def test_detects_dimensions(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        sales = next(v for v in discovery.variables if v.name == "Sales")
        assert "Geography" in sales.dimensions
        assert "Period" in sales.dimensions

    def test_computes_stats(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        sales = next(v for v in discovery.variables if v.name == "Sales")
        assert sales.stats.n_obs > 0
        assert sales.stats.mean > 0
        assert sales.stats.coverage_pct > 0

    def test_guesses_frequency(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        assert discovery.frequency_guess == "W"  # Weekly data

    def test_classifies_roles(self, sample_mff_df):
        discovery = discover_mff(sample_mff_df)
        sales = next(v for v in discovery.variables if v.name == "Sales")
        tv = next(v for v in discovery.variables if v.name == "TV_Spend")
        price = next(v for v in discovery.variables if v.name == "Price")

        assert sales.role == VariableRole.KPI
        assert tv.role == VariableRole.MEDIA
        assert price.role == VariableRole.CONTROL


# =============================================================================
# Template Generation Tests
# =============================================================================


class TestTemplateGenerator:
    """Tests for Excel template generation."""

    def test_generates_xlsx(self, sample_mff_df, tmp_dir):
        output_path = tmp_dir / "test_template.xlsx"
        result = TemplateGenerator.from_mff(sample_mff_df, output_path=output_path)

        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_has_all_sheets(self, sample_mff_df, tmp_dir):
        output_path = tmp_dir / "test_template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        assert "Variables" in wb.sheetnames
        assert "Media Settings" in wb.sheetnames
        assert "Model Settings" in wb.sheetnames
        assert "Advanced" in wb.sheetnames
        wb.close()

    def test_variables_sheet_has_data(self, sample_mff_df, tmp_dir):
        output_path = tmp_dir / "test_template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Variables"]

        # Should have variable names in the sheet
        values = [ws.cell(row=r, column=1).value for r in range(4, 12)]
        var_names = [v for v in values if v and v != "Variable Name"]
        assert len(var_names) == 7  # All 7 variables
        wb.close()

    def test_media_settings_populated(self, sample_mff_df, tmp_dir):
        output_path = tmp_dir / "test_template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Media Settings"]

        # Should have media channels listed
        values = [ws.cell(row=r, column=1).value for r in range(4, 10)]
        media_names = [v for v in values if v and v != "Variable Name"]
        assert len(media_names) == 3  # TV, Digital, Radio
        wb.close()

    def test_model_settings_defaults(self, sample_mff_df, tmp_dir):
        output_path = tmp_dir / "test_template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Model Settings"]

        # Find "Model Type" row and check default
        for row in ws.iter_rows(min_row=4, max_col=2):
            if row[0].value == "Model Type":
                assert row[1].value == "additive"
                break
        wb.close()

    def test_from_csv_file(self, sample_mff_df, tmp_dir):
        """Test generation from a CSV file path."""
        csv_path = tmp_dir / "data.csv"
        sample_mff_df.to_csv(csv_path, index=False)

        output_path = tmp_dir / "template.xlsx"
        result = TemplateGenerator.from_mff(csv_path, output_path=output_path)
        assert result.exists()

    def test_national_data_disables_geo_pooling(self, national_mff_df, tmp_dir):
        """When there are no geographies, geo pooling should default to False."""
        output_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(national_mff_df, output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        ws = wb["Model Settings"]

        for row in ws.iter_rows(min_row=4, max_col=2):
            if row[0].value == "Hierarchical Pooling (Geo)":
                assert row[1].value is False or row[1].value == "FALSE"
                break
        wb.close()


# =============================================================================
# Template Parser Tests
# =============================================================================


class TestTemplateParser:
    """Tests for parsing Excel templates back to configs."""

    def test_round_trip_basic(self, sample_mff_df, tmp_dir):
        """Generate template, parse it back, verify config structure."""
        template_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=template_path)

        mff_config, model_config, trend_config = TemplateParser.parse(template_path)

        # Check KPI
        assert mff_config.kpi.name == "Sales"

        # Check media channels
        media_names = mff_config.media_names
        assert "TV_Spend" in media_names
        assert "Digital_Spend" in media_names
        assert "Radio_Spend" in media_names
        assert len(media_names) == 3

        # Check controls
        control_names = mff_config.control_names
        assert "Price" in control_names

        # Check model config defaults
        assert model_config.specification == ModelSpecification.ADDITIVE
        assert model_config.inference_method == InferenceMethod.BAYESIAN_NUMPYRO
        assert model_config.n_chains == 4
        assert model_config.n_draws == 1000

    def test_round_trip_media_settings(self, sample_mff_df, tmp_dir):
        """Verify media channel settings survive round-trip."""
        template_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=template_path)

        mff_config, _, _ = TemplateParser.parse(template_path)

        for mc in mff_config.media_channels:
            assert mc.adstock.type == AdstockType.GEOMETRIC
            assert mc.adstock.l_max == 8
            assert mc.saturation.type == SaturationType.HILL

    def test_round_trip_dimensions(self, sample_mff_df, tmp_dir):
        """Verify dimensions survive round-trip."""
        template_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=template_path)

        mff_config, _, _ = TemplateParser.parse(template_path)

        # KPI should have Period and Geography dimensions
        assert DimensionType.PERIOD in mff_config.kpi.dimensions
        assert DimensionType.GEOGRAPHY in mff_config.kpi.dimensions

    def test_round_trip_geo_product(self, geo_product_mff_df, tmp_dir):
        """Test round-trip with both geo and product dimensions."""
        template_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(geo_product_mff_df, output_path=template_path)

        mff_config, model_config, _ = TemplateParser.parse(template_path)

        # Should have both geo and product dimensions
        assert DimensionType.GEOGRAPHY in mff_config.kpi.dimensions
        assert DimensionType.PRODUCT in mff_config.kpi.dimensions

        # Hierarchical pooling should be enabled for both
        assert model_config.hierarchical.pool_across_geo is True
        assert model_config.hierarchical.pool_across_product is True

    def test_file_not_found(self, tmp_dir):
        """Should raise error for non-existent file."""
        with pytest.raises(TemplateParseError, match="not found"):
            TemplateParser.parse(tmp_dir / "nonexistent.xlsx")

    def test_missing_variables_sheet(self, tmp_dir):
        """Should raise error when Variables sheet is missing."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "Wrong Sheet"
        path = tmp_dir / "bad_template.xlsx"
        wb.save(str(path))
        wb.close()

        with pytest.raises(TemplateParseError, match="Variables"):
            TemplateParser.parse(path)

    def test_no_kpi_error(self, tmp_dir):
        """Should raise validation error when no KPI is defined."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Variables"
        ws.cell(row=3, column=1, value="Variable Name")
        ws.cell(row=3, column=2, value="Role")
        ws.cell(row=3, column=3, value="Display Name")
        ws.cell(row=3, column=4, value="Dimensions")
        # Only media, no KPI
        ws.cell(row=4, column=1, value="TV_Spend")
        ws.cell(row=4, column=2, value="Media")
        ws.cell(row=4, column=3, value="TV")
        ws.cell(row=4, column=4, value="Period")

        path = tmp_dir / "no_kpi.xlsx"
        wb.save(str(path))
        wb.close()

        with pytest.raises(TemplateValidationError, match="No KPI"):
            TemplateParser.parse(path)

    def test_no_media_error(self, tmp_dir):
        """Should raise validation error when no media is defined."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Variables"
        ws.cell(row=3, column=1, value="Variable Name")
        ws.cell(row=3, column=2, value="Role")
        ws.cell(row=3, column=3, value="Display Name")
        ws.cell(row=3, column=4, value="Dimensions")
        # Only KPI, no media
        ws.cell(row=4, column=1, value="Sales")
        ws.cell(row=4, column=2, value="KPI")
        ws.cell(row=4, column=3, value="Sales")
        ws.cell(row=4, column=4, value="Period")

        path = tmp_dir / "no_media.xlsx"
        wb.save(str(path))
        wb.close()

        with pytest.raises(TemplateValidationError, match="No media"):
            TemplateParser.parse(path)

    def test_multiple_kpi_error(self, tmp_dir):
        """Should raise validation error for multiple KPIs."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Variables"
        ws.cell(row=3, column=1, value="Variable Name")
        ws.cell(row=3, column=2, value="Role")
        ws.cell(row=3, column=3, value="Display Name")
        ws.cell(row=3, column=4, value="Dimensions")
        # Two KPIs
        ws.cell(row=4, column=1, value="Sales")
        ws.cell(row=4, column=2, value="KPI")
        ws.cell(row=4, column=3, value="Sales")
        ws.cell(row=4, column=4, value="Period")
        ws.cell(row=5, column=1, value="Revenue")
        ws.cell(row=5, column=2, value="KPI")
        ws.cell(row=5, column=3, value="Revenue")
        ws.cell(row=5, column=4, value="Period")
        ws.cell(row=6, column=1, value="TV")
        ws.cell(row=6, column=2, value="Media")
        ws.cell(row=6, column=3, value="TV")
        ws.cell(row=6, column=4, value="Period")

        path = tmp_dir / "multi_kpi.xlsx"
        wb.save(str(path))
        wb.close()

        with pytest.raises(TemplateValidationError, match="Multiple KPI"):
            TemplateParser.parse(path)

    def test_frequency_detection(self, sample_mff_df, tmp_dir):
        """Frequency should be detected and round-tripped."""
        template_path = tmp_dir / "template.xlsx"
        TemplateGenerator.from_mff(sample_mff_df, output_path=template_path)

        mff_config, _, _ = TemplateParser.parse(template_path)
        assert mff_config.frequency == "W"
