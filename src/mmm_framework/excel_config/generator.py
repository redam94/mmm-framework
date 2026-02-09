"""
Excel config template generator.

Analyzes MFF data to discover variables, dimensions, and statistics,
then generates a pre-filled Excel workbook that analysts can edit
to configure a Marketing Mix Model.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook

from ..config import DimensionType, MFFColumnConfig, VariableRole
from .heuristics import VariableStats, classify_variable, generate_display_name
from .styles import (
    ROLE_FILLS,
    THIN_BORDER,
    WRAP_ALIGNMENT,
    Fills,
    Fonts,
    auto_fit_columns,
    create_adstock_validation,
    create_allocation_validation,
    create_boolean_validation,
    create_control_selection_validation,
    create_effect_direction_validation,
    create_frequency_validation,
    create_inference_validation,
    create_model_type_validation,
    create_prior_dist_validation,
    create_product_allocation_validation,
    create_role_validation,
    create_saturation_validation,
    create_trend_validation,
    format_header_row,
    write_instruction_row,
)


# =============================================================================
# Data classes for discovered variable info
# =============================================================================


@dataclass
class DiscoveredVariable:
    """Information about a variable discovered from MFF data."""

    name: str
    role: VariableRole
    display_name: str
    dimensions: list[str]  # Human-readable dimension names
    stats: VariableStats
    dimension_levels: dict[str, list[str]] = field(default_factory=dict)
    # e.g., {"Geography": ["North", "South", "West"], "Product": ["A", "B"]}


@dataclass
class MFFDiscovery:
    """Complete discovery results from analyzing MFF data."""

    variables: list[DiscoveredVariable]
    geographies: list[str]
    products: list[str]
    date_range: tuple[str, str]  # (min_date, max_date)
    n_periods: int
    frequency_guess: str  # "W", "D", or "M"
    column_config: MFFColumnConfig


# =============================================================================
# Discovery functions
# =============================================================================


def _read_data(data: pd.DataFrame | str | Path) -> pd.DataFrame:
    """Read MFF data from various sources."""
    if isinstance(data, pd.DataFrame):
        return data

    path = Path(data)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(path)
    elif suffix == ".parquet":
        return pd.read_parquet(path)
    elif suffix in (".xlsx", ".xls"):
        return pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use CSV, Parquet, or Excel.")


def _detect_column_config(df: pd.DataFrame) -> MFFColumnConfig:
    """
    Detect MFF column mappings from the DataFrame.

    Tries default names first, then fuzzy matches.
    """
    defaults = MFFColumnConfig()
    columns = set(df.columns)

    # Check if default column names exist
    if set(defaults.all_columns).issubset(columns):
        return defaults

    # Try case-insensitive matching
    col_map = {c.lower(): c for c in df.columns}

    def find_col(default_name: str) -> str:
        """Find the best matching column name."""
        if default_name in columns:
            return default_name
        lower = default_name.lower()
        if lower in col_map:
            return col_map[lower]
        # Try common variants
        variants = [
            lower.replace("_", ""),
            lower.replace("_", " "),
        ]
        for v in variants:
            if v in col_map:
                return col_map[v]
        return default_name  # Fall back to default

    return MFFColumnConfig(
        period=find_col("Period"),
        geography=find_col("Geography"),
        product=find_col("Product"),
        campaign=find_col("Campaign"),
        outlet=find_col("Outlet"),
        creative=find_col("Creative"),
        variable_name=find_col("VariableName"),
        variable_value=find_col("VariableValue"),
    )


def _detect_dimensions(
    df: pd.DataFrame,
    var_name: str,
    cols: MFFColumnConfig,
) -> tuple[list[str], dict[str, list[str]]]:
    """
    Detect which dimensions a variable varies across.

    Returns (dimension_names, dimension_levels) where dimension_levels
    maps dimension name → list of unique values for that dimension.
    """
    var_data = df[df[cols.variable_name] == var_name]

    if var_data.empty:
        return ["Period"], {}

    dimensions = ["Period"]
    levels: dict[str, list[str]] = {}

    dim_checks = [
        (cols.geography, "Geography"),
        (cols.product, "Product"),
        (cols.campaign, "Campaign"),
        (cols.outlet, "Outlet"),
        (cols.creative, "Creative"),
    ]

    for col, dim_name in dim_checks:
        if col in var_data.columns:
            unique_vals = var_data[col].dropna().unique()
            # Filter out empty strings and "Total"/"All" placeholders
            meaningful = [
                str(v) for v in unique_vals
                if str(v).strip() not in ("", "Total", "All", "National", "nan")
            ]
            if len(meaningful) > 1:
                dimensions.append(dim_name)
                levels[dim_name] = sorted(meaningful)

    return dimensions, levels


def _compute_stats(
    df: pd.DataFrame,
    var_name: str,
    cols: MFFColumnConfig,
    n_expected_combos: int,
) -> VariableStats:
    """Compute summary statistics for a variable."""
    var_data = df[df[cols.variable_name] == var_name]
    values = pd.to_numeric(var_data[cols.variable_value], errors="coerce").dropna()

    if values.empty:
        return VariableStats(
            mean=0.0,
            std=0.0,
            min_val=0.0,
            max_val=0.0,
            zero_pct=1.0,
            n_obs=0,
            coverage_pct=0.0,
        )

    n_obs = len(values)
    return VariableStats(
        mean=float(values.mean()),
        std=float(values.std()) if n_obs > 1 else 0.0,
        min_val=float(values.min()),
        max_val=float(values.max()),
        zero_pct=float((values == 0).sum() / n_obs) if n_obs > 0 else 0.0,
        n_obs=n_obs,
        coverage_pct=min(100.0, float(n_obs / max(n_expected_combos, 1) * 100)),
    )


def _guess_frequency(df: pd.DataFrame, cols: MFFColumnConfig) -> str:
    """Guess the data frequency from date differences."""
    try:
        dates = pd.to_datetime(df[cols.period].unique())
        dates = np.sort(dates)
        if len(dates) < 2:
            return "W"
        diffs = np.diff(dates)
        median_diff = np.median(diffs)
        days = median_diff / np.timedelta64(1, "D")
        if days <= 2:
            return "D"
        elif days <= 10:
            return "W"
        else:
            return "M"
    except Exception:
        return "W"


def discover_mff(
    data: pd.DataFrame | str | Path,
    column_config: MFFColumnConfig | None = None,
) -> MFFDiscovery:
    """
    Analyze MFF data and discover all variables, dimensions, and statistics.

    Parameters
    ----------
    data : DataFrame, str, or Path
        MFF data as a DataFrame or path to CSV/Parquet/Excel file.
    column_config : MFFColumnConfig, optional
        Column name mappings. Auto-detected if not provided.

    Returns
    -------
    MFFDiscovery
        Complete discovery results.
    """
    df = _read_data(data)
    cols = column_config or _detect_column_config(df)

    # Check required columns
    if cols.variable_name not in df.columns:
        raise ValueError(
            f"Column '{cols.variable_name}' not found in data. "
            f"Available columns: {list(df.columns)}"
        )
    if cols.variable_value not in df.columns:
        raise ValueError(
            f"Column '{cols.variable_value}' not found in data. "
            f"Available columns: {list(df.columns)}"
        )

    # Discover global info
    all_var_names = sorted(df[cols.variable_name].unique())

    # Get date range
    try:
        dates = pd.to_datetime(df[cols.period])
        date_range = (str(dates.min().date()), str(dates.max().date()))
        n_periods = int(dates.nunique())
    except Exception:
        date_range = ("unknown", "unknown")
        n_periods = 0

    # Get geo and product levels
    geographies: list[str] = []
    products: list[str] = []

    if cols.geography in df.columns:
        geo_vals = df[cols.geography].dropna().unique()
        geographies = sorted([
            str(v) for v in geo_vals
            if str(v).strip() not in ("", "Total", "All", "National", "nan")
        ])

    if cols.product in df.columns:
        prod_vals = df[cols.product].dropna().unique()
        products = sorted([
            str(v) for v in prod_vals
            if str(v).strip() not in ("", "Total", "All", "nan")
        ])

    frequency_guess = _guess_frequency(df, cols)

    # Estimate expected combos for coverage calculation
    n_expected = max(n_periods, 1) * max(len(geographies), 1) * max(len(products), 1)

    # Discover each variable
    variables: list[DiscoveredVariable] = []
    for var_name in all_var_names:
        dimensions, levels = _detect_dimensions(df, var_name, cols)
        stats = _compute_stats(df, var_name, cols, n_expected)
        role = classify_variable(var_name, stats)
        display = generate_display_name(var_name)

        variables.append(DiscoveredVariable(
            name=var_name,
            role=role,
            display_name=display,
            dimensions=dimensions,
            stats=stats,
            dimension_levels=levels,
        ))

    return MFFDiscovery(
        variables=variables,
        geographies=geographies,
        products=products,
        date_range=date_range,
        n_periods=n_periods,
        frequency_guess=frequency_guess,
        column_config=cols,
    )


# =============================================================================
# Template Generator
# =============================================================================


class TemplateGenerator:
    """
    Generate an Excel configuration template from MFF data.

    Usage
    -----
    >>> path = TemplateGenerator.from_mff("data.csv")
    >>> path = TemplateGenerator.from_mff(df, output_path="config.xlsx")
    """

    @classmethod
    def from_mff(
        cls,
        data: pd.DataFrame | str | Path,
        column_config: MFFColumnConfig | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """
        Analyze MFF data and generate a pre-filled Excel config template.

        Parameters
        ----------
        data : DataFrame, str, or Path
            MFF data as a DataFrame or path to a data file.
        column_config : MFFColumnConfig, optional
            Column name mappings. Auto-detected if not provided.
        output_path : str or Path, optional
            Where to save the Excel file. Defaults to "mmm_config_template.xlsx".

        Returns
        -------
        Path
            Path to the generated Excel file.
        """
        if output_path is None:
            output_path = Path("mmm_config_template.xlsx")
        else:
            output_path = Path(output_path)

        # Discover variables and dimensions
        discovery = discover_mff(data, column_config)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Variables
        ws_vars = wb.active
        ws_vars.title = "Variables"
        cls._write_variables_sheet(ws_vars, discovery)

        # Sheet 2: Media Settings
        ws_media = wb.create_sheet("Media Settings")
        media_vars = [v for v in discovery.variables if v.role == VariableRole.MEDIA]
        cls._write_media_settings_sheet(ws_media, media_vars)

        # Sheet 3: Model Settings
        ws_model = wb.create_sheet("Model Settings")
        cls._write_model_settings_sheet(ws_model, discovery)

        # Sheet 4: Advanced
        ws_advanced = wb.create_sheet("Advanced")
        control_vars = [v for v in discovery.variables if v.role == VariableRole.CONTROL]
        cls._write_advanced_sheet(ws_advanced, media_vars, control_vars)

        # Save
        wb.save(str(output_path))
        return output_path

    @classmethod
    def from_discovery(
        cls,
        discovery: MFFDiscovery,
        output_path: str | Path | None = None,
    ) -> Path:
        """
        Generate template from pre-computed discovery results.

        Useful when you want to modify discovery results before generating.
        """
        if output_path is None:
            output_path = Path("mmm_config_template.xlsx")
        else:
            output_path = Path(output_path)

        wb = Workbook()

        ws_vars = wb.active
        ws_vars.title = "Variables"
        cls._write_variables_sheet(ws_vars, discovery)

        ws_media = wb.create_sheet("Media Settings")
        media_vars = [v for v in discovery.variables if v.role == VariableRole.MEDIA]
        cls._write_media_settings_sheet(ws_media, media_vars)

        ws_model = wb.create_sheet("Model Settings")
        cls._write_model_settings_sheet(ws_model, discovery)

        ws_advanced = wb.create_sheet("Advanced")
        control_vars = [v for v in discovery.variables if v.role == VariableRole.CONTROL]
        cls._write_advanced_sheet(ws_advanced, media_vars, control_vars)

        wb.save(str(output_path))
        return output_path

    # =========================================================================
    # Sheet writers
    # =========================================================================

    @classmethod
    def _write_variables_sheet(
        cls,
        ws: Any,
        discovery: MFFDiscovery,
    ) -> None:
        """Write the Variables sheet with discovered variable info."""

        # -- Instructions --
        num_cols = 7
        instruction_text = (
            "VARIABLES CONFIGURATION\n"
            "Set the Role for each variable: KPI (target metric), Media (advertising channels), "
            "Control (non-media factors), or Exclude (ignore).\n"
            "Exactly ONE variable must be KPI. At least ONE must be Media.\n"
            f"Data: {discovery.n_periods} periods, "
            f"{len(discovery.geographies)} geographies, "
            f"{len(discovery.products)} products | "
            f"Date range: {discovery.date_range[0]} to {discovery.date_range[1]}"
        )
        write_instruction_row(ws, 1, instruction_text, num_cols)
        ws.row_dimensions[1].height = 60

        # -- Headers --
        headers = [
            "Variable Name",
            "Role",
            "Display Name",
            "Dimensions",
            "Data Coverage",
            "Mean Value",
            "Notes",
        ]
        header_row = 3
        format_header_row(ws, header_row, headers)

        # -- Data validation for Role column --
        role_dv = create_role_validation()
        ws.add_data_validation(role_dv)

        # -- Data rows --
        for i, var in enumerate(discovery.variables):
            row = header_row + 1 + i

            # Map role to display name
            role_display = {
                VariableRole.KPI: "KPI",
                VariableRole.MEDIA: "Media",
                VariableRole.CONTROL: "Control",
                VariableRole.AUXILIARY: "Exclude",
            }.get(var.role, "Exclude")

            ws.cell(row=row, column=1, value=var.name).font = Fonts.BOLD
            role_cell = ws.cell(row=row, column=2, value=role_display)
            ws.cell(row=row, column=3, value=var.display_name)
            ws.cell(row=row, column=4, value=", ".join(var.dimensions))
            ws.cell(row=row, column=5, value=f"{var.stats.coverage_pct:.1f}%")
            ws.cell(row=row, column=6, value=round(var.stats.mean, 2) if var.stats.mean != 0 else 0)
            ws.cell(row=row, column=7, value="")  # Notes

            # Apply role validation to the role cell
            role_dv.add(role_cell)

            # Color-code the row by role
            row_fill = ROLE_FILLS.get(role_display, Fills.EXCLUDE)
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = row_fill
                cell.border = THIN_BORDER

            # Mark read-only cells (stats columns)
            for col in [1, 4, 5, 6]:
                ws.cell(row=row, column=col).fill = Fills.LOCKED if role_display != "Exclude" else Fills.EXCLUDE

        auto_fit_columns(ws)
        # Ensure Role column is wide enough
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["G"].width = 30

    @classmethod
    def _write_media_settings_sheet(
        cls,
        ws: Any,
        media_vars: list[DiscoveredVariable],
    ) -> None:
        """Write the Media Settings sheet."""

        num_cols = 6
        instruction_text = (
            "MEDIA CHANNEL SETTINGS\n"
            "Configure adstock (carryover) and saturation for each media channel.\n"
            "Defaults are sensible starting points — adjust based on domain knowledge.\n"
            "Parent Channel: group related channels (e.g., 'Social' for Meta, TikTok, etc.)"
        )
        write_instruction_row(ws, 1, instruction_text, num_cols)
        ws.row_dimensions[1].height = 55

        headers = [
            "Variable Name",
            "Adstock Type",
            "Max Lag (weeks)",
            "Saturation Type",
            "Parent Channel",
            "Effect Direction",
        ]
        header_row = 3
        format_header_row(ws, header_row, headers)

        # Data validations
        adstock_dv = create_adstock_validation()
        saturation_dv = create_saturation_validation()
        direction_dv = create_effect_direction_validation()
        ws.add_data_validation(adstock_dv)
        ws.add_data_validation(saturation_dv)
        ws.add_data_validation(direction_dv)

        for i, var in enumerate(media_vars):
            row = header_row + 1 + i

            ws.cell(row=row, column=1, value=var.name).font = Fonts.BOLD
            adstock_cell = ws.cell(row=row, column=2, value="geometric")
            ws.cell(row=row, column=3, value=8)
            sat_cell = ws.cell(row=row, column=4, value="hill")
            ws.cell(row=row, column=5, value="")  # Parent channel
            dir_cell = ws.cell(row=row, column=6, value="positive")

            adstock_dv.add(adstock_cell)
            saturation_dv.add(sat_cell)
            direction_dv.add(dir_cell)

            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = Fills.MEDIA
                cell.border = THIN_BORDER

        if not media_vars:
            write_instruction_row(
                ws, header_row + 1,
                "No media variables detected. Set variable roles on the Variables sheet first, "
                "then add rows here manually.",
                num_cols,
            )

        auto_fit_columns(ws)
        ws.column_dimensions["E"].width = 20

    @classmethod
    def _write_model_settings_sheet(
        cls,
        ws: Any,
        discovery: MFFDiscovery,
    ) -> None:
        """Write the Model Settings sheet (key-value format)."""

        num_cols = 3
        instruction_text = (
            "MODEL SETTINGS\n"
            "Global model configuration. Adjust these settings to control "
            "model structure, inference, and hierarchical pooling.\n"
            "Defaults are recommended starting points."
        )
        write_instruction_row(ws, 1, instruction_text, num_cols)
        ws.row_dimensions[1].height = 50

        headers = ["Setting", "Value", "Description"]
        header_row = 3
        format_header_row(ws, header_row, headers)

        # Determine smart defaults based on discovery
        has_geo = len(discovery.geographies) > 1
        has_product = len(discovery.products) > 1

        settings: list[tuple[str, Any, str, Any]] = [
            # (key, default_value, description, validation_or_None)
            ("Model Type", "additive", "additive or multiplicative (log-transformed)", create_model_type_validation()),
            ("Inference Method", "bayesian_numpyro", "Sampling backend (numpyro is faster)", create_inference_validation()),
            ("Chains", 4, "Number of MCMC chains (≥2 recommended)", None),
            ("Draws", 1000, "Samples per chain after tuning", None),
            ("Tune", 1000, "Warmup/tuning samples per chain", None),
            ("Target Accept", 0.9, "Target acceptance rate (0.8-0.95)", None),
            ("Trend Type", "linear", "Baseline trend: none, linear, piecewise, spline", create_trend_validation()),
            ("Yearly Seasonality Order", 2, "Fourier order for yearly seasonality (0 = disabled)", None),
            ("Hierarchical Pooling (Geo)", has_geo, "Pool coefficients across geographies", create_boolean_validation()),
            ("Hierarchical Pooling (Product)", has_product, "Pool coefficients across products", create_boolean_validation()),
            ("Use Non-Centered", True, "Non-centered parameterization (better for sparse data)", create_boolean_validation()),
            ("Geo Allocation Method", "sales", "How to disaggregate national data to geos", create_allocation_validation()),
            ("Product Allocation Method", "sales", "How to disaggregate data to products", create_product_allocation_validation()),
            ("Data Frequency", discovery.frequency_guess, "Data granularity: W(eekly), D(aily), M(onthly)", create_frequency_validation()),
            ("Control Selection", "none", "Automatic variable selection: none, horseshoe, spike_slab", create_control_selection_validation()),
        ]

        for i, (key, default, description, dv) in enumerate(settings):
            row = header_row + 1 + i

            ws.cell(row=row, column=1, value=key).font = Fonts.BOLD
            val_cell = ws.cell(row=row, column=2, value=default)
            ws.cell(row=row, column=3, value=description).font = Fonts.INSTRUCTION

            if dv is not None:
                ws.add_data_validation(dv)
                dv.add(val_cell)

            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

        # Add data summary section
        summary_row = header_row + len(settings) + 2
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=num_cols)
        cell = ws.cell(row=summary_row, column=1, value="DATA SUMMARY (Read-Only)")
        cell.font = Fonts.BOLD
        cell.fill = Fills.LOCKED

        summary_items = [
            ("Date Range", f"{discovery.date_range[0]} to {discovery.date_range[1]}"),
            ("Number of Periods", discovery.n_periods),
            ("Geographies", ", ".join(discovery.geographies) if discovery.geographies else "None (national)"),
            ("Products", ", ".join(discovery.products) if discovery.products else "None (single product)"),
            ("Detected Frequency", discovery.frequency_guess),
        ]

        for j, (label, value) in enumerate(summary_items):
            row = summary_row + 1 + j
            ws.cell(row=row, column=1, value=label).font = Fonts.BOLD
            ws.cell(row=row, column=2, value=value)
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = Fills.LOCKED
                ws.cell(row=row, column=col).border = THIN_BORDER

        auto_fit_columns(ws, min_width=15)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 55

    @classmethod
    def _write_advanced_sheet(
        cls,
        ws: Any,
        media_vars: list[DiscoveredVariable],
        control_vars: list[DiscoveredVariable],
    ) -> None:
        """Write the Advanced sheet with prior override options."""

        num_cols = 11
        instruction_text = (
            "ADVANCED SETTINGS (Optional)\n"
            "Override default priors and transformation parameters for fine-grained control.\n"
            "Leave defaults unless you have specific domain knowledge about parameter values.\n"
            "All priors use sensible defaults if not modified here."
        )
        write_instruction_row(ws, 1, instruction_text, num_cols)
        ws.row_dimensions[1].height = 55

        # -- Section A: Media Priors --
        section_a_row = 3
        ws.merge_cells(start_row=section_a_row, start_column=1, end_row=section_a_row, end_column=num_cols)
        cell = ws.cell(row=section_a_row, column=1, value="MEDIA CHANNEL PRIORS")
        cell.font = Fonts.BOLD
        cell.fill = Fills.MEDIA

        media_headers = [
            "Variable Name",
            "Coefficient Prior",
            "Coefficient σ",
            "Adstock α Prior",
            "Adstock α a",
            "Adstock α b",
            "Saturation κ Low",
            "Saturation κ High",
            "Slope Prior",
            "Slope α",
            "Slope β",
        ]
        media_header_row = section_a_row + 1
        format_header_row(ws, media_header_row, media_headers)

        # Validations for prior distributions
        coeff_dv = create_prior_dist_validation("half_normal,gamma,log_normal")
        adstock_dv = create_prior_dist_validation("beta")
        slope_dv = create_prior_dist_validation("gamma,half_normal,log_normal")
        ws.add_data_validation(coeff_dv)
        ws.add_data_validation(adstock_dv)
        ws.add_data_validation(slope_dv)

        for i, var in enumerate(media_vars):
            row = media_header_row + 1 + i

            ws.cell(row=row, column=1, value=var.name).font = Fonts.BOLD
            coeff_cell = ws.cell(row=row, column=2, value="half_normal")
            ws.cell(row=row, column=3, value=2.0)
            adstock_cell = ws.cell(row=row, column=4, value="beta")
            ws.cell(row=row, column=5, value=1.0)
            ws.cell(row=row, column=6, value=3.0)
            ws.cell(row=row, column=7, value=0.1)
            ws.cell(row=row, column=8, value=0.9)
            slope_cell = ws.cell(row=row, column=9, value="gamma")
            ws.cell(row=row, column=10, value=2.0)
            ws.cell(row=row, column=11, value=1.0)

            coeff_dv.add(coeff_cell)
            adstock_dv.add(adstock_cell)
            slope_dv.add(slope_cell)

            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

        # -- Section B: Control Priors --
        section_b_row = media_header_row + len(media_vars) + 3
        ws.merge_cells(start_row=section_b_row, start_column=1, end_row=section_b_row, end_column=num_cols)
        cell = ws.cell(row=section_b_row, column=1, value="CONTROL VARIABLE PRIORS")
        cell.font = Fonts.BOLD
        cell.fill = Fills.CONTROL

        control_headers = [
            "Variable Name",
            "Allow Negative",
            "Coefficient Prior",
            "Coefficient μ",
            "Coefficient σ",
            "Use Shrinkage",
        ]
        control_header_row = section_b_row + 1
        format_header_row(ws, control_header_row, control_headers[:6])

        ctrl_prior_dv = create_prior_dist_validation("normal,half_normal")
        bool_dv = create_boolean_validation()
        ws.add_data_validation(ctrl_prior_dv)
        ws.add_data_validation(bool_dv)

        for i, var in enumerate(control_vars):
            row = control_header_row + 1 + i

            # Guess allow_negative based on variable name
            allow_neg = "TRUE"
            name_lower = var.name.lower()
            if any(kw in name_lower for kw in ("price", "discount", "promo")):
                allow_neg = "TRUE"  # Price can go up or down
            elif any(kw in name_lower for kw in ("distribution", "store", "temperature")):
                allow_neg = "TRUE"

            ws.cell(row=row, column=1, value=var.name).font = Fonts.BOLD
            neg_cell = ws.cell(row=row, column=2, value=allow_neg)
            prior_cell = ws.cell(row=row, column=3, value="normal")
            ws.cell(row=row, column=4, value=0.0)
            ws.cell(row=row, column=5, value=1.0)
            shrink_cell = ws.cell(row=row, column=6, value="FALSE")

            bool_dv.add(neg_cell)
            ctrl_prior_dv.add(prior_cell)
            bool_dv.add(shrink_cell)

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER

        auto_fit_columns(ws, min_width=14)
