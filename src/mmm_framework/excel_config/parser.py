"""
Excel config template parser.

Reads a filled-in Excel configuration template and constructs
MFFConfig + ModelConfig objects ready for model fitting.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import (
    AdstockConfig,
    AdstockType,
    AllocationMethod,
    ControlSelectionConfig,
    ControlVariableConfig,
    DimensionAlignmentConfig,
    DimensionType,
    HierarchicalConfig,
    InferenceMethod,
    KPIConfig,
    MediaChannelConfig,
    MFFColumnConfig,
    MFFConfig,
    ModelConfig,
    ModelSpecification,
    PriorConfig,
    PriorType,
    SaturationConfig,
    SaturationType,
    SeasonalityConfig,
)
from ..model.trend_config import TrendConfig, TrendType


# =============================================================================
# Exceptions
# =============================================================================


class TemplateParseError(Exception):
    """Raised when a template cannot be parsed."""

    pass


class TemplateValidationError(Exception):
    """Raised when a parsed template has invalid configuration."""

    pass


# =============================================================================
# Helper functions
# =============================================================================


def _to_bool(value: Any) -> bool:
    """Convert an Excel cell value to a Python bool."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().upper() in ("TRUE", "YES", "1", "Y")
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _to_float(value: Any, default: float = 0.0) -> float:
    """Convert an Excel cell value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _to_int(value: Any, default: int = 0) -> int:
    """Convert an Excel cell value to int."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def _to_str(value: Any, default: str = "") -> str:
    """Convert an Excel cell value to string."""
    if value is None:
        return default
    return str(value).strip()


def _parse_dimensions(dim_str: str) -> list[DimensionType]:
    """Parse a comma-separated dimension string into DimensionType list."""
    if not dim_str:
        return [DimensionType.PERIOD]

    dim_map = {
        "period": DimensionType.PERIOD,
        "geography": DimensionType.GEOGRAPHY,
        "geo": DimensionType.GEOGRAPHY,
        "product": DimensionType.PRODUCT,
        "campaign": DimensionType.CAMPAIGN,
        "outlet": DimensionType.OUTLET,
        "creative": DimensionType.CREATIVE,
    }

    dims = []
    for part in dim_str.split(","):
        part = part.strip().lower()
        if part in dim_map:
            dims.append(dim_map[part])

    if not dims or DimensionType.PERIOD not in dims:
        dims.insert(0, DimensionType.PERIOD)

    return dims


def _build_prior(
    dist_name: str,
    params: dict[str, float],
) -> PriorConfig:
    """Build a PriorConfig from distribution name and parameters."""
    dist_map = {
        "half_normal": PriorType.HALF_NORMAL,
        "normal": PriorType.NORMAL,
        "log_normal": PriorType.LOG_NORMAL,
        "gamma": PriorType.GAMMA,
        "beta": PriorType.BETA,
        "truncated_normal": PriorType.TRUNCATED_NORMAL,
        "half_student_t": PriorType.HALF_STUDENT_T,
    }

    dist_type = dist_map.get(dist_name.lower(), PriorType.HALF_NORMAL)
    return PriorConfig(distribution=dist_type, params=params)


# =============================================================================
# Sheet parsers
# =============================================================================


def _parse_variables_sheet(ws: Any) -> list[dict[str, Any]]:
    """
    Parse the Variables sheet.

    Returns a list of dicts with keys: name, role, display_name, dimensions.
    Skips instruction and header rows.
    """
    variables = []

    # Find the header row (contains "Variable Name")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=7):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Variable Name":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        raise TemplateParseError(
            "Could not find 'Variable Name' header in Variables sheet. "
            "Make sure the sheet has the expected structure."
        )

    # Read data rows starting after header
    for row in ws.iter_rows(min_row=header_row + 1, max_col=7):
        name = _to_str(row[0].value)
        role = _to_str(row[1].value)
        display_name = _to_str(row[2].value)
        dimensions = _to_str(row[3].value)

        if not name or not role:
            continue  # Skip empty rows

        if role.lower() == "exclude":
            continue  # Skip excluded variables

        variables.append({
            "name": name,
            "role": role.lower(),
            "display_name": display_name or name,
            "dimensions": dimensions,
        })

    return variables


def _parse_media_settings_sheet(ws: Any) -> dict[str, dict[str, Any]]:
    """
    Parse the Media Settings sheet.

    Returns a dict mapping variable_name → settings dict.
    """
    settings: dict[str, dict[str, Any]] = {}

    # Find header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=6):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Variable Name":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        return settings  # Empty media settings is OK — will use defaults

    for row in ws.iter_rows(min_row=header_row + 1, max_col=6):
        name = _to_str(row[0].value)
        if not name:
            continue

        settings[name] = {
            "adstock_type": _to_str(row[1].value, "geometric"),
            "max_lag": _to_int(row[2].value, 8),
            "saturation_type": _to_str(row[3].value, "hill"),
            "parent_channel": _to_str(row[4].value) or None,
            "effect_direction": _to_str(row[5].value, "positive"),
        }

    return settings


def _parse_model_settings_sheet(ws: Any) -> dict[str, Any]:
    """
    Parse the Model Settings sheet (key-value format).

    Returns a dict mapping setting_name → value.
    """
    settings: dict[str, Any] = {}

    # Find header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=3):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Setting":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        return settings

    for row in ws.iter_rows(min_row=header_row + 1, max_col=2):
        key = _to_str(row[0].value)
        value = row[1].value

        if not key:
            continue
        # Stop at the data summary section
        if key.upper().startswith("DATA SUMMARY"):
            break

        settings[key] = value

    return settings


def _parse_advanced_sheet(ws: Any) -> tuple[dict[str, dict], dict[str, dict]]:
    """
    Parse the Advanced sheet.

    Returns (media_priors, control_priors) where each maps
    variable_name → prior settings dict.
    """
    media_priors: dict[str, dict] = {}
    control_priors: dict[str, dict] = {}

    current_section = None
    header_row = None

    for row in ws.iter_rows(min_row=1, max_col=11):
        first_val = _to_str(row[0].value)

        # Detect section headers
        if "MEDIA CHANNEL PRIORS" in first_val.upper():
            current_section = "media"
            header_row = None
            continue
        elif "CONTROL VARIABLE PRIORS" in first_val.upper():
            current_section = "control"
            header_row = None
            continue

        # Detect column headers within sections
        if first_val == "Variable Name":
            header_row = row[0].row
            continue

        if header_row is None or not first_val:
            continue

        if current_section == "media":
            name = first_val
            media_priors[name] = {
                "coefficient_dist": _to_str(row[1].value, "half_normal"),
                "coefficient_sigma": _to_float(row[2].value, 2.0),
                "adstock_dist": _to_str(row[3].value, "beta"),
                "adstock_a": _to_float(row[4].value, 1.0),
                "adstock_b": _to_float(row[5].value, 3.0),
                "kappa_low": _to_float(row[6].value, 0.1),
                "kappa_high": _to_float(row[7].value, 0.9),
                "slope_dist": _to_str(row[8].value, "gamma"),
                "slope_a": _to_float(row[9].value, 2.0),
                "slope_b": _to_float(row[10].value, 1.0),
            }

        elif current_section == "control":
            name = first_val
            control_priors[name] = {
                "allow_negative": _to_bool(row[1].value) if row[1].value is not None else True,
                "coefficient_dist": _to_str(row[2].value, "normal"),
                "coefficient_mu": _to_float(row[3].value, 0.0),
                "coefficient_sigma": _to_float(row[4].value, 1.0),
                "use_shrinkage": _to_bool(row[5].value) if row[5].value is not None else False,
            }

    return media_priors, control_priors


# =============================================================================
# Config builders
# =============================================================================


def _build_media_channel_config(
    var_info: dict[str, Any],
    media_settings: dict[str, Any] | None,
    media_priors: dict[str, Any] | None,
) -> MediaChannelConfig:
    """Build a MediaChannelConfig from parsed sheet data."""

    name = var_info["name"]
    dimensions = _parse_dimensions(var_info["dimensions"])

    # Media settings (from Sheet 2)
    ms = media_settings or {}
    adstock_type_str = ms.get("adstock_type", "geometric")
    max_lag = ms.get("max_lag", 8)
    saturation_type_str = ms.get("saturation_type", "hill")
    parent_channel = ms.get("parent_channel")
    effect_direction = ms.get("effect_direction", "positive")

    # Advanced priors (from Sheet 4)
    mp = media_priors or {}

    # Build adstock config
    adstock_type = AdstockType(adstock_type_str)
    if adstock_type == AdstockType.NONE:
        adstock = AdstockConfig.none()
    else:
        alpha_prior = _build_prior(
            mp.get("adstock_dist", "beta"),
            {"alpha": mp.get("adstock_a", 1.0), "beta": mp.get("adstock_b", 3.0)},
        )
        adstock = AdstockConfig(
            type=adstock_type,
            l_max=max_lag,
            alpha_prior=alpha_prior,
        )

    # Build saturation config
    sat_type = SaturationType(saturation_type_str)
    if sat_type == SaturationType.NONE:
        saturation = SaturationConfig.none()
    else:
        slope_prior = _build_prior(
            mp.get("slope_dist", "gamma"),
            {"alpha": mp.get("slope_a", 2.0), "beta": mp.get("slope_b", 1.0)},
        )
        saturation = SaturationConfig(
            type=sat_type,
            slope_prior=slope_prior,
            kappa_bounds_percentiles=(
                mp.get("kappa_low", 0.1),
                mp.get("kappa_high", 0.9),
            ),
        )

    # Build coefficient prior
    if effect_direction == "positive":
        coeff_prior = _build_prior(
            mp.get("coefficient_dist", "half_normal"),
            {"sigma": mp.get("coefficient_sigma", 2.0)},
        )
    else:
        coeff_prior = _build_prior("normal", {"mu": 0.0, "sigma": mp.get("coefficient_sigma", 2.0)})

    return MediaChannelConfig(
        name=name,
        display_name=var_info.get("display_name", name),
        dimensions=dimensions,
        adstock=adstock,
        saturation=saturation,
        coefficient_prior=coeff_prior,
        parent_channel=parent_channel,
    )


def _build_control_config(
    var_info: dict[str, Any],
    control_priors: dict[str, Any] | None,
) -> ControlVariableConfig:
    """Build a ControlVariableConfig from parsed sheet data."""

    name = var_info["name"]
    dimensions = _parse_dimensions(var_info["dimensions"])

    cp = control_priors or {}
    allow_negative = cp.get("allow_negative", True)
    use_shrinkage = cp.get("use_shrinkage", False)

    if allow_negative:
        coeff_prior = _build_prior(
            cp.get("coefficient_dist", "normal"),
            {"mu": cp.get("coefficient_mu", 0.0), "sigma": cp.get("coefficient_sigma", 1.0)},
        )
    else:
        coeff_prior = _build_prior(
            "half_normal",
            {"sigma": cp.get("coefficient_sigma", 1.0)},
        )

    return ControlVariableConfig(
        name=name,
        display_name=var_info.get("display_name", name),
        dimensions=dimensions,
        allow_negative=allow_negative,
        coefficient_prior=coeff_prior,
        use_shrinkage=use_shrinkage,
    )


def _build_kpi_config(
    var_info: dict[str, Any],
    model_settings: dict[str, Any],
) -> KPIConfig:
    """Build a KPIConfig from parsed sheet data."""

    name = var_info["name"]
    dimensions = _parse_dimensions(var_info["dimensions"])
    log_transform = _to_str(model_settings.get("Model Type", "additive")).lower() == "multiplicative"

    return KPIConfig(
        name=name,
        display_name=var_info.get("display_name", name),
        dimensions=dimensions,
        log_transform=log_transform,
    )


def _build_model_config(model_settings: dict[str, Any]) -> ModelConfig:
    """Build a ModelConfig from the Model Settings sheet."""

    # Model specification
    spec_str = _to_str(model_settings.get("Model Type", "additive")).lower()
    specification = (
        ModelSpecification.MULTIPLICATIVE
        if spec_str == "multiplicative"
        else ModelSpecification.ADDITIVE
    )

    # Inference method
    inf_str = _to_str(model_settings.get("Inference Method", "bayesian_numpyro")).lower()
    inf_map = {
        "bayesian_numpyro": InferenceMethod.BAYESIAN_NUMPYRO,
        "bayesian_pymc": InferenceMethod.BAYESIAN_PYMC,
        "frequentist_ridge": InferenceMethod.FREQUENTIST_RIDGE,
    }
    inference = inf_map.get(inf_str, InferenceMethod.BAYESIAN_NUMPYRO)

    # MCMC settings
    n_chains = _to_int(model_settings.get("Chains", 4), 4)
    n_draws = _to_int(model_settings.get("Draws", 1000), 1000)
    n_tune = _to_int(model_settings.get("Tune", 1000), 1000)
    target_accept = _to_float(model_settings.get("Target Accept", 0.9), 0.9)

    # Seasonality
    yearly_order = _to_int(model_settings.get("Yearly Seasonality Order", 2), 2)
    seasonality = SeasonalityConfig(
        yearly=yearly_order if yearly_order > 0 else None,
    )

    # Hierarchical
    pool_geo = _to_bool(model_settings.get("Hierarchical Pooling (Geo)", True))
    pool_product = _to_bool(model_settings.get("Hierarchical Pooling (Product)", True))
    use_nc = _to_bool(model_settings.get("Use Non-Centered", True))

    hierarchical = HierarchicalConfig(
        enabled=pool_geo or pool_product,
        pool_across_geo=pool_geo,
        pool_across_product=pool_product,
        use_non_centered=use_nc,
    )

    # Control selection
    cs_method = _to_str(model_settings.get("Control Selection", "none")).lower()
    control_selection = ControlSelectionConfig(method=cs_method)

    return ModelConfig(
        specification=specification,
        inference_method=inference,
        n_chains=n_chains,
        n_draws=n_draws,
        n_tune=n_tune,
        target_accept=target_accept,
        seasonality=seasonality,
        hierarchical=hierarchical,
        control_selection=control_selection,
    )


def _build_trend_config(model_settings: dict[str, Any]) -> TrendConfig:
    """Build a TrendConfig from Model Settings."""
    trend_str = _to_str(model_settings.get("Trend Type", "linear")).lower()
    trend_map = {
        "none": TrendType.NONE,
        "linear": TrendType.LINEAR,
        "piecewise": TrendType.PIECEWISE,
        "spline": TrendType.SPLINE,
    }
    trend_type = trend_map.get(trend_str, TrendType.LINEAR)
    return TrendConfig(type=trend_type)


def _build_dimension_alignment(model_settings: dict[str, Any]) -> DimensionAlignmentConfig:
    """Build DimensionAlignmentConfig from Model Settings."""
    geo_alloc_str = _to_str(model_settings.get("Geo Allocation Method", "sales")).lower()
    prod_alloc_str = _to_str(model_settings.get("Product Allocation Method", "sales")).lower()

    alloc_map = {
        "equal": AllocationMethod.EQUAL,
        "population": AllocationMethod.POPULATION,
        "sales": AllocationMethod.SALES,
        "custom": AllocationMethod.CUSTOM,
    }

    return DimensionAlignmentConfig(
        geo_allocation=alloc_map.get(geo_alloc_str, AllocationMethod.SALES),
        product_allocation=alloc_map.get(prod_alloc_str, AllocationMethod.SALES),
    )


# =============================================================================
# Main Parser
# =============================================================================


class TemplateParser:
    """
    Parse a filled-in Excel config template into framework configuration objects.

    Usage
    -----
    >>> mff_config, model_config, trend_config = TemplateParser.parse("config.xlsx")
    >>> loader = MFFLoader(mff_config)
    >>> panel = loader.load(data)
    """

    @classmethod
    def parse(
        cls,
        template_path: str | Path,
    ) -> tuple[MFFConfig, ModelConfig, TrendConfig]:
        """
        Parse an Excel configuration template.

        Parameters
        ----------
        template_path : str or Path
            Path to the filled-in Excel template.

        Returns
        -------
        tuple[MFFConfig, ModelConfig, TrendConfig]
            Ready-to-use configuration objects.

        Raises
        ------
        TemplateParseError
            If the template structure is invalid.
        TemplateValidationError
            If the parsed configuration is logically invalid.
        """
        path = Path(template_path)
        if not path.exists():
            raise TemplateParseError(f"Template file not found: {path}")

        wb = load_workbook(str(path), data_only=True)

        # Parse all sheets
        if "Variables" not in wb.sheetnames:
            raise TemplateParseError("Template must have a 'Variables' sheet.")

        variables = _parse_variables_sheet(wb["Variables"])

        media_settings = {}
        if "Media Settings" in wb.sheetnames:
            media_settings = _parse_media_settings_sheet(wb["Media Settings"])

        model_settings = {}
        if "Model Settings" in wb.sheetnames:
            model_settings = _parse_model_settings_sheet(wb["Model Settings"])

        media_priors: dict[str, dict] = {}
        control_priors: dict[str, dict] = {}
        if "Advanced" in wb.sheetnames:
            media_priors, control_priors = _parse_advanced_sheet(wb["Advanced"])

        wb.close()

        # Validate: must have exactly 1 KPI
        kpi_vars = [v for v in variables if v["role"] == "kpi"]
        if len(kpi_vars) == 0:
            raise TemplateValidationError(
                "No KPI variable defined. Set exactly one variable's Role to 'KPI'."
            )
        if len(kpi_vars) > 1:
            raise TemplateValidationError(
                f"Multiple KPI variables defined: {[v['name'] for v in kpi_vars]}. "
                "Set exactly one variable's Role to 'KPI'."
            )

        # Validate: must have at least 1 media
        media_vars = [v for v in variables if v["role"] == "media"]
        if len(media_vars) == 0:
            raise TemplateValidationError(
                "No media variables defined. Set at least one variable's Role to 'Media'."
            )

        control_vars = [v for v in variables if v["role"] == "control"]

        # Build configs
        kpi_config = _build_kpi_config(kpi_vars[0], model_settings)

        media_configs = [
            _build_media_channel_config(
                v,
                media_settings.get(v["name"]),
                media_priors.get(v["name"]),
            )
            for v in media_vars
        ]

        control_configs = [
            _build_control_config(v, control_priors.get(v["name"]))
            for v in control_vars
        ]

        # Data frequency
        freq = _to_str(model_settings.get("Data Frequency", "W"), "W").upper()
        if freq not in ("W", "D", "M"):
            freq = "W"

        alignment = _build_dimension_alignment(model_settings)

        mff_config = MFFConfig(
            kpi=kpi_config,
            media_channels=media_configs,
            controls=control_configs,
            alignment=alignment,
            frequency=freq,
        )

        model_config = _build_model_config(model_settings)
        trend_config = _build_trend_config(model_settings)

        return mff_config, model_config, trend_config
