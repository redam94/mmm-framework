"""
Excel formatting constants and utilities for config templates.

Provides consistent styling, data validation rules, and
helper functions for openpyxl workbook formatting.
"""

from __future__ import annotations

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# Color Palette (matching MMM Framework design tokens)
# =============================================================================

class Colors:
    """Color constants for Excel formatting."""

    # Role colors (fill backgrounds)
    KPI_FILL = "E8F5E9"        # Light green
    MEDIA_FILL = "E3F2FD"      # Light blue
    CONTROL_FILL = "FFF3E0"    # Light orange
    EXCLUDE_FILL = "F5F5F5"    # Light gray

    # Header colors
    HEADER_FILL = "4A6D4A"     # Dark green (matches --color-primary-dark)
    HEADER_FONT = "FFFFFF"     # White text
    SUBHEADER_FILL = "8FA86A"  # Green (matches --color-primary)
    SUBHEADER_FONT = "FFFFFF"

    # Instruction row
    INSTRUCTION_FILL = "FFFDE7"  # Light yellow
    INSTRUCTION_FONT = "5A6B5A"  # Muted green

    # Data validation
    LOCKED_FILL = "ECEFF1"     # Light blue-gray (read-only cells)

    # Borders
    BORDER_COLOR = "D4DDD4"    # Matches --color-border


# =============================================================================
# Font Definitions
# =============================================================================

class Fonts:
    """Font presets for Excel formatting."""

    HEADER = Font(
        name="Calibri",
        size=11,
        bold=True,
        color=Colors.HEADER_FONT,
    )
    SUBHEADER = Font(
        name="Calibri",
        size=10,
        bold=True,
        color=Colors.SUBHEADER_FONT,
    )
    INSTRUCTION = Font(
        name="Calibri",
        size=9,
        italic=True,
        color=Colors.INSTRUCTION_FONT,
    )
    NORMAL = Font(
        name="Calibri",
        size=10,
    )
    BOLD = Font(
        name="Calibri",
        size=10,
        bold=True,
    )
    SHEET_TITLE = Font(
        name="Calibri",
        size=14,
        bold=True,
        color="2D3A2D",  # --color-text
    )


# =============================================================================
# Fill Definitions
# =============================================================================

class Fills:
    """Fill presets for Excel formatting."""

    HEADER = PatternFill(start_color=Colors.HEADER_FILL, end_color=Colors.HEADER_FILL, fill_type="solid")
    SUBHEADER = PatternFill(start_color=Colors.SUBHEADER_FILL, end_color=Colors.SUBHEADER_FILL, fill_type="solid")
    INSTRUCTION = PatternFill(start_color=Colors.INSTRUCTION_FILL, end_color=Colors.INSTRUCTION_FILL, fill_type="solid")
    LOCKED = PatternFill(start_color=Colors.LOCKED_FILL, end_color=Colors.LOCKED_FILL, fill_type="solid")

    KPI = PatternFill(start_color=Colors.KPI_FILL, end_color=Colors.KPI_FILL, fill_type="solid")
    MEDIA = PatternFill(start_color=Colors.MEDIA_FILL, end_color=Colors.MEDIA_FILL, fill_type="solid")
    CONTROL = PatternFill(start_color=Colors.CONTROL_FILL, end_color=Colors.CONTROL_FILL, fill_type="solid")
    EXCLUDE = PatternFill(start_color=Colors.EXCLUDE_FILL, end_color=Colors.EXCLUDE_FILL, fill_type="solid")


# =============================================================================
# Border & Alignment
# =============================================================================

THIN_BORDER = Border(
    left=Side(style="thin", color=Colors.BORDER_COLOR),
    right=Side(style="thin", color=Colors.BORDER_COLOR),
    top=Side(style="thin", color=Colors.BORDER_COLOR),
    bottom=Side(style="thin", color=Colors.BORDER_COLOR),
)

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


# =============================================================================
# Role-to-fill mapping
# =============================================================================

ROLE_FILLS = {
    "KPI": Fills.KPI,
    "Media": Fills.MEDIA,
    "Control": Fills.CONTROL,
    "Exclude": Fills.EXCLUDE,
}


# =============================================================================
# Data Validation Helpers
# =============================================================================

def create_role_validation() -> DataValidation:
    """Create dropdown validation for variable roles."""
    dv = DataValidation(
        type="list",
        formula1='"KPI,Media,Control,Exclude"',
        allow_blank=False,
    )
    dv.error = "Please select: KPI, Media, Control, or Exclude"
    dv.errorTitle = "Invalid Role"
    dv.prompt = "Select the variable's role in the model"
    dv.promptTitle = "Variable Role"
    return dv


def create_adstock_validation() -> DataValidation:
    """Create dropdown validation for adstock types."""
    dv = DataValidation(
        type="list",
        formula1='"geometric,weibull,delayed,none"',
        allow_blank=False,
    )
    dv.error = "Please select: geometric, weibull, delayed, or none"
    dv.errorTitle = "Invalid Adstock Type"
    return dv


def create_saturation_validation() -> DataValidation:
    """Create dropdown validation for saturation types."""
    dv = DataValidation(
        type="list",
        formula1='"hill,logistic,michaelis_menten,tanh,none"',
        allow_blank=False,
    )
    dv.error = "Please select: hill, logistic, michaelis_menten, tanh, or none"
    dv.errorTitle = "Invalid Saturation Type"
    return dv


def create_effect_direction_validation() -> DataValidation:
    """Create dropdown validation for effect direction."""
    dv = DataValidation(
        type="list",
        formula1='"positive,any"',
        allow_blank=False,
    )
    dv.error = "Please select: positive or any"
    dv.errorTitle = "Invalid Direction"
    return dv


def create_model_type_validation() -> DataValidation:
    """Create dropdown for additive vs multiplicative."""
    return DataValidation(type="list", formula1='"additive,multiplicative"', allow_blank=False)


def create_inference_validation() -> DataValidation:
    """Create dropdown for inference method."""
    return DataValidation(type="list", formula1='"bayesian_numpyro,bayesian_pymc"', allow_blank=False)


def create_trend_validation() -> DataValidation:
    """Create dropdown for trend type."""
    return DataValidation(type="list", formula1='"none,linear,piecewise,spline"', allow_blank=False)


def create_frequency_validation() -> DataValidation:
    """Create dropdown for data frequency."""
    return DataValidation(type="list", formula1='"W,D,M"', allow_blank=False)


def create_allocation_validation() -> DataValidation:
    """Create dropdown for allocation method."""
    return DataValidation(type="list", formula1='"equal,population,sales"', allow_blank=False)


def create_product_allocation_validation() -> DataValidation:
    """Create dropdown for product allocation method."""
    return DataValidation(type="list", formula1='"equal,sales"', allow_blank=False)


def create_control_selection_validation() -> DataValidation:
    """Create dropdown for control variable selection method."""
    return DataValidation(type="list", formula1='"none,horseshoe,spike_slab"', allow_blank=False)


def create_boolean_validation() -> DataValidation:
    """Create dropdown for boolean values."""
    return DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)


def create_prior_dist_validation(distributions: str = "half_normal,gamma,log_normal,normal") -> DataValidation:
    """Create dropdown for prior distribution selection."""
    return DataValidation(type="list", formula1=f'"{distributions}"', allow_blank=False)


# =============================================================================
# Sheet formatting helpers
# =============================================================================

def format_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Apply header formatting to a row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Fonts.HEADER
        cell.fill = Fills.HEADER
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def format_subheader_row(ws: Worksheet, row: int, text: str, num_cols: int) -> None:
    """Apply subheader/section formatting to a merged row."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=num_cols,
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Fonts.SUBHEADER
    cell.fill = Fills.SUBHEADER
    cell.alignment = CENTER_ALIGNMENT


def write_instruction_row(ws: Worksheet, row: int, text: str, num_cols: int) -> None:
    """Write an instruction row spanning multiple columns."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=num_cols,
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Fonts.INSTRUCTION
    cell.fill = Fills.INSTRUCTION
    cell.alignment = WRAP_ALIGNMENT


def auto_fit_columns(ws: Worksheet, min_width: int = 12, max_width: int = 40) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width
